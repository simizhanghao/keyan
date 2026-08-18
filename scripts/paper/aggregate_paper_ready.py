#!/usr/bin/env python3
"""Aggregate all paper_runs into outputs/paper_ready/."""

from __future__ import annotations

import argparse
import csv
import json
import shutil
from pathlib import Path

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


def load_metrics(run_dir: Path) -> list[dict]:
    rows = []
    for metrics in run_dir.rglob("metrics.json"):
        rel = metrics.relative_to(run_dir)
        m = json.loads(metrics.read_text())
        m["experiment"] = str(rel.parent)
        m["metrics_path"] = str(metrics)
        rows.append(m)
    return rows


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    keys = sorted({k for r in rows for k in r})
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", default=".")
    args = parser.parse_args()
    root = Path(args.root)
    ready = root / "outputs" / "paper_ready"
    ready.mkdir(parents=True, exist_ok=True)
    figs = ready / "figures"
    figs.mkdir(exist_ok=True)

    runs_root = root / "outputs" / "paper_runs"
    all_metrics = []
    for phase_dir in sorted(runs_root.glob("phase*")):
        all_metrics.extend(load_metrics(phase_dir))

    write_csv(ready / "all_metrics_summary.csv", all_metrics)

    # Copy phase-specific summaries if present
    for name in [
        "ablation_summary.csv",
        "deployment_shift_summary.csv",
        "cross_receiver_summary.csv",
        "edge_deployment_summary.csv",
    ]:
        for p in runs_root.rglob(name):
            shutil.copy2(p, ready / name)
            break

    # Bootstrap / paired from phase2 stats
    for pattern in ["*bootstrap*.csv", "*paired*.csv"]:
        for p in runs_root.rglob(pattern):
            dest = ready / p.name
            if not dest.exists():
                shutil.copy2(p, dest)

    # manifest audit
    audit = root / "outputs" / "paper_ready" / "manifest_audit.csv"
    if not audit.exists():
        import subprocess
        subprocess.run([__import__("sys").executable, str(root / "scripts/paper/audit_manifests.py")], cwd=root)

    # XLSX workbook
    if HAS_XLSX:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for csv_name in sorted(ready.glob("*.csv")):
            ws = wb.create_sheet(csv_name.stem[:31])
            with csv_name.open(encoding="utf-8") as f:
                for i, row in enumerate(csv.reader(f)):
                    for j, val in enumerate(row):
                        ws.cell(row=i + 1, column=j + 1, value=val)
        wb.save(ready / "main_tables.xlsx")

    commit = "unknown"
    try:
        import subprocess
        commit = subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=root, text=True).strip()
    except Exception:
        pass

    readme = f"""# Paper Results Package

Generated from `outputs/paper_runs/` aggregation.

Git commit: `{commit}`

## Files

| File | Description |
|------|-------------|
| all_metrics_summary.csv | All metrics.json merged |
| manifest_audit.csv | Phase 1 manifest audit |
| ablation_summary.csv | Phase 3 component ablation |
| deployment_shift_summary.csv | Phase 4 Config/Location/Distance |
| cross_receiver_summary.csv | Phase 5 cross-receiver |
| edge_deployment_summary.csv | Phase 6 latency/params |
| main_tables.xlsx | Excel workbook of all CSVs |

## Scripts

| Phase | Script |
|-------|--------|
| 1 | scripts/paper/phase1_manifests.sh |
| 2 | scripts/paper/phase2_cross_day.sh |
| 3 | scripts/paper/phase3_ablation.sh |
| 4 | scripts/paper/phase4_deployment.sh |
| 5 | scripts/paper/phase5_cross_receiver.sh |
| 6 | scripts/paper/phase6_edge_benchmark.py |

## Protocol

- **Source-only (paper main):** train/val from source domain, test on target. See `data/paper/*_source_only.csv`.
- **Oracle diagnostic:** target-val checkpoint selection. See `data/paper/*_oracle_target_val.csv`.

Do not overwrite old runs; each phase creates timestamped dir under `outputs/paper_runs/`.
"""
    (ready / "README_paper_results.md").write_text(readme, encoding="utf-8")
    print(f"paper_ready updated: {ready}")


if __name__ == "__main__":
    main()
